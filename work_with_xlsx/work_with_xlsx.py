from re import I
import openpyxl

inventory_file = openpyxl.load_workbook("work_with_xml/inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 10:
        products_under_10_inv[product_num] = inventory

print(products_under_10_inv)
print(products_per_supplier)
print(total_value_per_supplier)